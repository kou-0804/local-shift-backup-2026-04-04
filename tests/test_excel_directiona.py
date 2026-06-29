# tests/test_excel_directiona.py
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries
from shift_scheduler.src.excel_directiona import render_directiona

STATS = ['夜勤', '病院MR', 'CLMR', '病CT', 'CT', 'ア', '心', 'ク', 'ポ', '精',
         'MG', 'DR', 'HB', 'OP', '入', '病L', '超遅', 'ク遅', 'M遅', '公休', '代休']


def _sample_grid():
    # 2026-06: day 6 = Sat(土), day 7 = Sun(日). Hand-built to exercise every kind
    # without invoking the solver.
    weekdays = {1: '月', 2: '火', 3: '水', 4: '木', 5: '金', 6: '土', 7: '日'}

    def meta(kind, fill=None):
        return {"kind": kind, "fill": fill}

    row = {
        "staff_id": "T001", "staff_num": 1, "name": "佐藤(海)",
        "cells":     {1: "CT", 2: "病CT夜", 3: "○", 4: "休", 5: "☆", 6: "", 7: "希望休"},
        "cell_meta": {1: meta("work"), 2: meta("night", "FFFF00"), 3: meta("akemei", "FFC0CB"),
                      4: meta("off", "D3D3D3"), 5: meta("special_off", "FFCDD2"),
                      6: meta("empty"), 7: meta("request")},
        "has_work": True,
        "stats": {k: 0 for k in STATS} | {"夜勤": 1, "CT": 1, "病CT": 1, "公休": 8},
    }
    return {
        "year": 2026, "month": 6, "days_in_month": 7,
        "weekdays": weekdays, "stats_columns": STATS,
        "rows": [row],
        "oncall_rows": [{"label": "第1拘束",
                         "cells": {d: ("佐藤海" if d == 1 else "") for d in range(1, 8)}}],
    }


def _load():
    return openpyxl.load_workbook(BytesIO(render_directiona(_sample_grid())))


def test_main_sheet_present_and_titled():
    wb = _load()
    assert wb.sheetnames[0].startswith("勤務")        # main grid sheet first


def test_title_merge_width_tracks_days_plus_offset():
    ws = _load().worksheets[0]
    # offset = 2 name/number cols + len(stats); width must equal days + offset (NOT 34).
    days, offset = 7, 2 + len(STATS)
    title_ranges = [r for r in ws.merged_cells.ranges if str(r).startswith("A1")]
    assert len(title_ranges) == 1
    min_c, _, max_c, _ = range_boundaries(str(title_ranges[0]))
    assert (max_c - min_c + 1) == days + offset


def test_two_row_header_dates_then_weekdays():
    ws = _load().worksheets[0]
    # row2 = dates 1..7 starting col C; row3 = weekday chars.
    assert [ws.cell(row=2, column=2 + d).value for d in range(1, 8)] == [1, 2, 3, 4, 5, 6, 7]
    assert [ws.cell(row=3, column=2 + d).value for d in range(1, 8)] == \
           ['月', '火', '水', '木', '金', '土', '日']


def test_weekend_and_holiday_columns_are_shaded():
    from tests.golden.xlsx_dump import fill_hex
    ws = _load().worksheets[0]
    sat = ws.cell(row=2, column=2 + 6)   # day 6 = 土
    sun = ws.cell(row=2, column=2 + 7)   # day 7 = 日
    assert fill_hex(sat) == "DDEBF7"     # 薄青
    assert fill_hex(sun) == "FCE4D6"     # 薄赤


def test_freeze_panes_locks_name_cols_and_header_rows():
    ws = _load().worksheets[0]
    assert ws.freeze_panes == "C4"


def test_night_cell_is_navy_with_white_font():
    from tests.golden.xlsx_dump import fill_hex
    ws = _load().worksheets[0]
    c = ws.cell(row=4, column=2 + 2)    # T001 day2 = 病CT夜 (kind=night)
    assert c.value == "病CT夜"
    assert fill_hex(c) == "1F3864"
    assert (c.font.color.rgb or "")[-6:] == "FFFFFF"


def test_akemei_grey_off_green_request_italic():
    from tests.golden.xlsx_dump import fill_hex
    ws = _load().worksheets[0]
    assert fill_hex(ws.cell(row=4, column=2 + 3)) == "D9D9D9"   # ○  akemei
    assert fill_hex(ws.cell(row=4, column=2 + 4)) == "E2EFDA"   # 休 off
    assert fill_hex(ws.cell(row=4, column=2 + 5)) == "E2EFDA"   # ☆ special_off
    assert ws.cell(row=4, column=2 + 7).font.italic is True     # 希望休 request


def test_oncall_row_has_thick_duty_border():
    ws = _load().worksheets[0]
    # one row of staff (row4) then the 第1拘束 row at row5
    label = ws.cell(row=5, column=1).value
    assert label == "第1拘束"
    cell = ws.cell(row=5, column=2 + 1)            # 佐藤海 on day 1
    assert cell.value == "佐藤海"
    assert cell.border.left.style == "medium"


def test_stats_block_written_for_working_row():
    ws = _load().worksheets[0]
    # 公休 is the 20th of 21 stats labels -> column = 2 + days + 20
    col = 2 + 7 + (STATS.index("公休") + 1)
    assert ws.cell(row=4, column=col).value == 8


def test_legend_and_summary_sheets_exist():
    wb = _load()
    assert "凡例" in wb.sheetnames
    assert "集計" in wb.sheetnames


def test_legend_explains_symbols_and_colours():
    ws = _load()["凡例"]
    blob = "\n".join(str(c.value) for col in ws.iter_cols() for c in col if c.value)
    for token in ("夜", "明", "(希)", "★", "☆", "17休", "公休"):
        assert token in blob


def test_summary_lists_each_staff_with_key_counts():
    ws = _load()["集計"]
    header = [c.value for c in ws[1]]
    for col in ("技師名", "公休", "夜勤", "代休"):
        assert col in header
    names = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    assert "佐藤(海)" in names


def test_summary_includes_per_location_columns():
    ws = _load()["集計"]
    header = [c.value for c in ws[1]]
    for loc in ("病CT", "CT", "CLMR"):     # 各場所 from stats_columns
        assert loc in header
