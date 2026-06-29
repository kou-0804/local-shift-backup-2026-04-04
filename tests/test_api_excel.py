# tests/test_api_excel.py
from io import BytesIO
from types import SimpleNamespace

import openpyxl
from fastapi.testclient import TestClient

from shift_scheduler.src.models.schedule_result import ScheduleResult
import webapp.api.main as api_main
from webapp.api.db import connect, init_db
from webapp.api.main import app
from webapp.api.rosters import freeze_roster


def _seed(conn):
    result = ScheduleResult(
        year=2026, month=6, staff=[{"id": "T001", "name": "佐藤(海)"}],
        day_assignments={1: {"CT": ["T001"]}, 2: {"ク": ["T001"]}},
        night_assignments={}, requests={}, on_call_assignments={},
        daikyu_counts={}, off_counts={"T001": 9}, validation_errors=[],
        daily_location_needs={2: {"ク": 1}})
    techs = [SimpleNamespace(id="T001", name="佐藤(海)", status="在籍",
                             note="", night_hb=False)]
    return freeze_roster(conn, job_id="j", result=result, technicians=techs,
                         data_dir="d", target_holidays=9)


def test_rosters_excel_returns_direction_a_workbook(tmp_path):
    conn = connect(str(tmp_path / "t.db"))
    init_db(conn)
    rid = _seed(conn)
    conn.commit()
    app.dependency_overrides[api_main.get_db] = lambda: conn
    try:
        client = TestClient(app)
        resp = client.get(f"/rosters/{rid}/excel")
        assert resp.status_code == 200
        assert resp.headers["content-type"] == \
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "attachment" in resp.headers["content-disposition"]
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        assert wb.worksheets[0].title.startswith("勤務")     # Direction-A main sheet
        assert "凡例" in wb.sheetnames and "集計" in wb.sheetnames
        assert wb.worksheets[0].freeze_panes == "C4"         # Direction-A, not legacy
    finally:
        app.dependency_overrides.clear()


def test_rosters_excel_404_for_unknown(tmp_path):
    conn = connect(str(tmp_path / "t.db"))
    init_db(conn)
    app.dependency_overrides[api_main.get_db] = lambda: conn
    try:
        client = TestClient(app)
        assert client.get("/rosters/999999/excel").status_code == 404
    finally:
        app.dependency_overrides.clear()
