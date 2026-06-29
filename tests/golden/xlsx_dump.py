"""Normalize an .xlsx to a JSON-able dict so two workbooks can be compared for
'byte-identity' independent of openpyxl's zip ordering. Shared by the parity
fixture generator and the parity test."""
from io import BytesIO
import openpyxl


def fill_hex(cell):
    """Solid-fill foreground colour as a 6-char hex, or None when unfilled.
    Normalizes openpyxl's optional 8-char ARGB ('00FFFF00') down to 'FFFF00'."""
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return None
    rgb = getattr(fill.fgColor, "rgb", None)
    if not isinstance(rgb, str):
        return None
    return rgb[-6:].upper()


def dump_workbook(xlsx_bytes) -> dict:
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
    sheets = {}
    for ws in wb.worksheets:
        cells = {}
        for row in ws.iter_rows():
            for c in row:
                if c.value is None and fill_hex(c) is None:
                    continue
                cells[c.coordinate] = [c.value, fill_hex(c)]
        sheets[ws.title] = {
            "cells": cells,
            "merged": sorted(str(r) for r in ws.merged_cells.ranges),
            "freeze_panes": ws.freeze_panes,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
        }
    return {"sheetnames": wb.sheetnames, "sheets": sheets}
