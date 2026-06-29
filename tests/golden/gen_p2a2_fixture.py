"""One-time generator for the P2a-2 build_grid golden. Reuses the assignments
already solved in tests/golden/2026-06_p2a1.json (NO solver run): it builds the
CURRENT ExcelGenerator, renders the xlsx, and parses the 21 stat columns,
has-work blanking, and on-call rows per active staff as ground truth that the
new build_grid must reproduce byte-for-byte.
Usage: python tests/golden/gen_p2a2_fixture.py
"""
import json
from io import BytesIO

import openpyxl

from shift_scheduler.src.loaders.data_loader import DataLoader
from shift_scheduler.src.excel_generator import ExcelGenerator

YEAR, MONTH, DATA_DIR = 2026, 6, "shift_scheduler/data"
P2A1 = "tests/golden/2026-06_p2a1.json"


def main():
    fix = json.load(open(P2A1, encoding="utf-8"))
    day = {int(d): v for d, v in fix["day_assignments"].items()}
    night = {int(d): v for d, v in fix["night_assignments"].items()}
    req = {int(d): v for d, v in fix["requests"].items()}
    off = fix["expected_off_counts"]
    daikyu = fix["expected_daikyu_counts"]

    technicians = DataLoader(data_dir=DATA_DIR).load_all(f"{YEAR}-{MONTH:02d}")[0]
    active = [t for t in technicians if t.status == "在籍"]

    gen = ExcelGenerator(
        year=YEAR, month=MONTH, technicians=technicians,
        night_assignments=night, day_assignments=day, requests=req,
        on_call_assignments={}, daikyu_counts=daikyu, off_counts=off,
        validation_errors=[],
    )
    wb = openpyxl.load_workbook(BytesIO(gen.generate_bytes()))
    ws = wb["6月勤務表"]  # main sheet (see tests/test_excel_bytes.py)

    stats_columns = gen.stats_columns
    days = gen.days_in_month
    # Stats begin at excel col `days_in_month + 3` (A=勤務表番号, B=技師名,
    # day cols span 3..days+2, stats start at days+3 == stats_start_col in
    # excel_generator.py:201). NB: the plan draft wrote `1 + 1 + days` which is
    # the LAST day column (off-by-one); corrected here to match the real source.
    first_stat_col = days + 3
    # Map staff_num -> row by scanning column A (勤務表番号).
    num_to_row = {}
    for r in range(4, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v not in (None, ""):
            num_to_row[str(v)] = r

    expected_stats, expected_has_work = {}, {}
    for t in active:
        try:
            num = str(int(t.id.replace("T", "")))
        except ValueError:
            num = t.id
        row = num_to_row.get(num)
        if row is None:
            expected_stats[t.id] = None
            expected_has_work[t.id] = False
            continue
        vals = [ws.cell(row=row, column=first_stat_col + i).value
                for i in range(len(stats_columns))]
        if all(v in (None, "") for v in vals):
            expected_stats[t.id] = None
            expected_has_work[t.id] = False
        else:
            expected_stats[t.id] = {
                lbl: (vals[i] if vals[i] not in (None, "") else 0)
                for i, lbl in enumerate(stats_columns)
            }
            expected_has_work[t.id] = True

    fixture = {
        "year": YEAR, "month": MONTH, "days_in_month": days,
        "stats_columns": stats_columns,
        "expected_stats": expected_stats,
        "expected_has_work": expected_has_work,
    }
    with open("tests/golden/2026-06_p2a2.json", "w", encoding="utf-8") as f:
        json.dump(fixture, f, ensure_ascii=False, sort_keys=True, indent=2)
    print(f"wrote tests/golden/2026-06_p2a2.json: {len(active)} active staff")


if __name__ == "__main__":
    main()
