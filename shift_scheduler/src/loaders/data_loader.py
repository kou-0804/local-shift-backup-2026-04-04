from typing import Tuple, List, Dict
import os
from .staff_loader import StaffLoader
from .skill_loader import SkillLoader
from .location_loader import LocationLoader
from .request_loader import RequestLoader
from .rule_loader import RuleLoader
from ..models.staff import Staff
from ..models.location import Location
from ..models.request import Request
from ..models.skill import SkillRank
from ..models.rule import SpecialRule
from ..models.power_balance import PowerBalance
from .power_balance_loader import PowerBalanceLoader

class DataLoader:
    def __init__(self, data_dir: str):
        self.data_dir = data_dir
        
    def load_all(self, month: str) -> Tuple[List[Staff], List[Location], Dict[str, Dict[str, SkillRank]], List[Request], List[SpecialRule], List[PowerBalance]]:
        
        # 1. Load Staff
        staff_path = os.path.join(self.data_dir, "技師マスタ_確定版.csv")
        staff_loader = StaffLoader(staff_path)
        staff_list = staff_loader.load()
        staff_ids = [s.id for s in staff_list]
        
        # 2. Load Skills
        skill_path = os.path.join(self.data_dir, "スキルマスタ_確定版.csv")
        skill_loader = SkillLoader(skill_path)
        skills = skill_loader.load(staff_ids)
        
        # 3. Derive Night Skills (Base from Skill Master)
        for staff in staff_list:
            s_skills = skills.get(staff.id, {})
            
            mr_rank = max(s_skills.get('病院MR', SkillRank.NONE), s_skills.get('クMR', SkillRank.NONE))
            staff.night_mr = (mr_rank >= SkillRank.B)
            
            angio_rank = s_skills.get('ア', SkillRank.NONE)
            staff.night_angio = (angio_rank >= SkillRank.B)
            
            cath_rank = s_skills.get('心', SkillRank.NONE)
            staff.night_cath = (cath_rank >= SkillRank.B)

        # 4. Apply Overrides from Night Skill List
        night_skill_path = os.path.join(self.data_dir, "夜勤スキル一覧.csv")
        if os.path.exists(night_skill_path):
            print(f"Loading Night Skill Overrides from {night_skill_path}")
            import pandas as pd
            try:
                ns_df = pd.read_csv(night_skill_path)
                ns_df.columns = [c.strip() for c in ns_df.columns]
                
                name_map = {s.name: s for s in staff_list}
                
                for _, row in ns_df.iterrows():
                    name = str(row['SName']).strip().replace('　', ' ')
                    staff = name_map.get(name)
                    if not staff and ' ' in name:
                         staff = name_map.get(name.replace(' ', '　'))
                    
                    if staff:
                        def get_bool_override(val):
                            if pd.isna(val) or val == '': return None
                            s_val = str(val).strip().upper()
                            if s_val == 'TRUE': return True
                            if s_val == 'FALSE': return False
                            return None

                        override_mr = get_bool_override(row.get('NightShiftMR'))
                        if override_mr is not None: staff.night_mr = override_mr
                        
                        override_cath = get_bool_override(row.get('NightShiftCardiacCath'))
                        if override_cath is not None: staff.night_cath = override_cath
                        
                        override_angio = get_bool_override(row.get('NightShiftAngio'))
                        if override_angio is not None: staff.night_angio = override_angio

            except Exception as e:
                print(f"Error reading Night Skills CSV: {e}")

        # 4. Load Locations
        loc_path = os.path.join(self.data_dir, "勤務場所マスタ_確定版.csv")
        location_loader = LocationLoader(loc_path)
        locations = location_loader.load()
        
        # 5. Load Special Rules
        rule_path = os.path.join(self.data_dir, "特殊配置ルール_確定版.csv")
        rule_loader = RuleLoader(rule_path)
        rules = rule_loader.load()
        
        # 6. Load Requests
        name_to_id = {s.name: s.id for s in staff_list}
        request_loader = RequestLoader(self.data_dir)
        requests = request_loader.load(month, name_to_id=name_to_id)
        
        # 7. Load Power Balance Rules
        # They are in the same file as locations
        pb_loader = PowerBalanceLoader(loc_path)
        pb_rules = pb_loader.load()
        
        return staff_list, locations, skills, requests, rules, pb_rules

    def load_night_counts(self, month: str, name_to_id: Dict[str, str]) -> Dict[str, int]:
        """月ごとの夜勤回数上限をロード (ファイル名: 夜勤回数_確定版.csv)"""
        import pandas as pd
        import unicodedata
        path = os.path.join(self.data_dir, "夜勤回数_確定版.csv")
        if not os.path.exists(path):
            print(f"Warning: Night count file not found: {path}")
            return {}
        
        try:
            # Skip first row which is just a title "1月夜勤回数"
            df = pd.read_csv(path, encoding='utf-8', skiprows=1)
            
            # Find target column for month
            m_int = int(month.split('-')[1])
            target_col = None
            
            # Normalize column names to handle full-width chars (１月 -> 1月)
            norm_cols = {col: unicodedata.normalize('NFKC', str(col)) for col in df.columns}
            
            # Look for "1月" or "1月夜勤回数"
            candidates = [f"{m_int}月", f"{m_int}月夜勤回数"]
            
            for col, norm_col in norm_cols.items():
                if norm_col in candidates:
                    target_col = col
                    break
            
            if not target_col:
                 print(f"Warning: Month column '{month}' (or {m_int}月) not found in {path}.")
                 # Fallback to the first column that looks like a month or just the second column
                 # The first column is usually '名前', the second is the month data e.g. '3月'
                 available_cols = list(df.columns)
                 if len(available_cols) > 1:
                     target_col = available_cols[1]
                     print(f"Fallback: Using column '{target_col}' for quotas instead.")
                 else:
                     return {}
            
            # Map Name -> ID -> Count
            result = {}
            for _, row in df.iterrows():
                raw_name = str(row['名前'])
                name = raw_name.strip().replace('　', ' ') # Normalize spaces
                
                sid = name_to_id.get(name)
                if not sid:
                    if ' ' in name:
                        fw_name = name.replace(' ', '　')
                        sid = name_to_id.get(fw_name)
                
                
                if not sid:
                    # Skip rows like "合計", "必要当直者数"
                    continue
                    
                val = row[target_col]
                # Handle empty or non-numeric
                if pd.isna(val) or val == '':
                    continue
                try:
                    result[sid] = int(val)
                except:
                    pass
            return result
        except Exception as e:
            print(f"Error loading night counts: {e}")
            return {}
    def load_night_history_from_excel(self, file_path: str, name_to_id: Dict[str, str]) -> List[Tuple[int, str]]:
        """
        指定された勤務表Excelから夜勤実績（日付day, staff_id）のリストを読み込む
        :param file_path: Excelファイルのパス
        :param name_to_id: 名前からStaff IDへのマッピング
        :return: List of (day, staff_id)
        """
        import openpyxl
        import unicodedata

        if not os.path.exists(file_path):
            print(f"  Note: Previous month file not found: {file_path}")
            return []
            
        print(f"  Loading history from {file_path}...")
        results = []
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Identify columns
            # Row 2 contains days "01", "02"...
            # Column B is "技師名"
            
            date_col_map = {} # day_int -> col_idx (1-based)
            staff_name_col = 2 # Column B
            
            header_row = 2
            for col in range(3, ws.max_column + 1):
                val = ws.cell(row=header_row, column=col).value
                if isinstance(val, int) or (isinstance(val, str) and val.isdigit()):
                    day_int = int(val)
                    date_col_map[day_int] = col
            
            # Iterate staff rows
            start_row = 4
            for row in range(start_row, ws.max_row + 1):
                name_cell = ws.cell(row=row, column=staff_name_col).value
                if not name_cell:
                    continue
                
                name = str(name_cell).strip().replace('　', ' ')
                sid = name_to_id.get(name)
                # Try fallback
                if not sid and ' ' in name:
                     sid = name_to_id.get(name.replace(' ', '　'))
                
                if not sid:
                    continue
                    
                # Check each day
                for day, col in date_col_map.items():
                    val = ws.cell(row=row, column=col).value
                    if val and isinstance(val, str):
                        # Check for Night symbol
                        # Usually "MR夜", "夜", "CT/夜" etc.
                        if '夜' in val:
                            results.append((day, sid))
                            
            return results
        except Exception as e:
            print(f"Error loading Excel history: {e}")
            return []
