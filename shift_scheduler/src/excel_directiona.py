"""Direction-A Excel renderer (spec §6.5). Thin openpyxl placement layer over the
build_grid dict: ZERO derivation. Distinct from the legacy ExcelGenerator — colours
key off cell_meta['kind'] and the layout is redesigned. Deterministic.

Holiday shading is single-sourced via stats_engine._is_public_off (no second
jpholiday code path)."""
from io import BytesIO
from datetime import date

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from shift_scheduler.src import stats_engine

# Direction-A palette (intentionally different from legacy value-driven fills)
C_NIGHT = "1F3864"    # 夜勤  濃紺 (white font)
C_AKEMEI = "D9D9D9"   # 明け  薄グレー
C_OFF = "E2EFDA"      # 公休  薄緑 (off / special_off)
C_SAT = "DDEBF7"      # 土    薄青
C_SUNHOL = "FCE4D6"   # 日/祝 薄赤
C_HEADER = "4472C4"   # ヘッダ

_NAME_COLS = 2         # A=勤務表番号, B=技師名
_FIRST_DATA_ROW = 4    # row1 title, row2 dates, row3 weekdays
_DUTY_BORDER = Border(*(Side(style="medium"),) * 4)   # 拘束=太枠


def _fill(hex_):
    return PatternFill(start_color=hex_, end_color=hex_, fill_type="solid")


def _day_col(day):
    return _NAME_COLS + day            # C = day 1


def _is_holiday(y, m, d):
    """日/祝 判定。stats_engine._is_public_off へ委譲（祝日ロジック単一ソース化）。"""
    return stats_engine._is_public_off(date(y, m, d))


def render_directiona(grid: dict, *, warnings=None) -> bytes:
    y, m, days = grid["year"], grid["month"], grid["days_in_month"]
    stats_cols = grid["stats_columns"]
    last_col = _NAME_COLS + days + len(stats_cols)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"勤務分担表{m}月"
    center = Alignment(horizontal="center", vertical="center")

    # Row 1: title, merged to the table width (tracks days; no AH1 hardcode).
    ws.cell(row=1, column=1, value=f"画像診断室 {y}年{m}月 勤務分担表").font = Font(size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.cell(row=1, column=1).alignment = center

    # Row 2/3 fixed labels + per-day date/weekday + weekend/holiday column shading.
    ws.cell(row=2, column=1, value="勤務表番号")
    ws.cell(row=2, column=2, value="技師名")
    ws.cell(row=3, column=2, value="曜日")
    for d in range(1, days + 1):
        col = _day_col(d)
        wd = grid["weekdays"][d]
        dc = ws.cell(row=2, column=col, value=d)
        dc.alignment = center
        wc = ws.cell(row=3, column=col, value=wd)
        wc.alignment = center
        shade = None
        if wd == "土":
            shade = C_SAT
        elif wd == "日" or _is_holiday(y, m, d):
            shade = C_SUNHOL
        if shade:
            dc.fill = wc.fill = _fill(shade)
    # Stats header labels start right after the day block.
    for i, label in enumerate(stats_cols):
        ws.cell(row=2, column=_NAME_COLS + days + 1 + i, value=label).alignment = center

    _render_body(ws, grid)            # Task 5
    _build_legend_sheet(wb)           # Task 6
    _build_summary_sheet(wb, grid)    # Task 6
    _build_validation_sheet(wb, y, m, warnings)   # Task 7
    _apply_print_setup(ws, last_col)  # Task 7

    ws.freeze_panes = ws.cell(row=_FIRST_DATA_ROW, column=_NAME_COLS + 1).coordinate  # 'C4'

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- Task 5/6/7 helpers (filled in by later tasks) ---


def _style_for(kind):
    """(fill_or_None, font_or_None) for a body cell, keyed off kind (Direction-A)."""
    if kind == "night":
        return _fill(C_NIGHT), Font(color="FFFFFF")
    if kind == "akemei":
        return _fill(C_AKEMEI), None
    if kind in ("off", "special_off"):
        return _fill(C_OFF), None
    if kind == "request":
        return None, Font(italic=True)          # 希望休 = 斜体
    return None, None                            # work / empty = 白


def _render_body(ws, grid):
    center = Alignment(horizontal="center", vertical="center")
    days = grid["days_in_month"]
    stats_cols = grid["stats_columns"]
    r = _FIRST_DATA_ROW
    for row in grid["rows"]:
        ws.cell(row=r, column=1, value=row["staff_num"])
        ws.cell(row=r, column=2, value=row["name"])
        for d in range(1, days + 1):
            text = row["cells"].get(d, "")
            kind = row["cell_meta"].get(d, {}).get("kind", "empty")
            c = ws.cell(row=r, column=_day_col(d), value=text)
            c.alignment = center
            fill, font = _style_for(kind)
            if fill:
                c.fill = fill
            if font:
                c.font = font
        if row["has_work"] and row["stats"]:
            for i, label in enumerate(stats_cols):
                ws.cell(row=r, column=_NAME_COLS + days + 1 + i,
                        value=row["stats"][label]).alignment = center
        r += 1
    # On-call rows with the 拘束 太枠.
    for oc in grid["oncall_rows"]:
        ws.cell(row=r, column=1, value=oc["label"])
        for d in range(1, days + 1):
            c = ws.cell(row=r, column=_day_col(d), value=oc["cells"].get(d, ""))
            c.alignment = center
            c.border = _DUTY_BORDER
        r += 1


def _build_legend_sheet(wb):
    """記号 + Direction-A パレットの意味を静的に列挙。"""
    ws = wb.create_sheet("凡例")
    ws.cell(row=1, column=1, value="凡例").font = Font(size=14, bold=True)
    rows = [
        ("夜", "夜勤（濃紺・白字）", C_NIGHT),
        ("明 / ○", "明け（薄グレー）", C_AKEMEI),
        ("公休 / 休", "公休（薄緑）", C_OFF),
        ("(希) / 希望休", "希望休（斜体・白）", None),
        ("★ / ☆ / ◆", "特別休（★連・☆小・☆デ 等）", None),
        ("17休", "17時以降休（時短）", None),
        ("土", "土曜（薄青）", C_SAT),
        ("日 / 祝", "日曜・祝日（薄赤）", C_SUNHOL),
    ]
    r = 3
    for sym, desc, swatch in rows:
        ws.cell(row=r, column=1, value=sym)
        ws.cell(row=r, column=2, value=desc)
        if swatch:
            ws.cell(row=r, column=3).fill = _fill(swatch)
        r += 1
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 40


def _build_summary_sheet(wb, grid):
    """個人別集計シート: 勤務表番号・技師名 + 各 stats_columns（公休/夜勤/代休/各場所）。"""
    ws = wb.create_sheet("集計")
    center = Alignment(horizontal="center", vertical="center")
    stats_cols = grid["stats_columns"]
    header = ["勤務表番号", "技師名"] + list(stats_cols)
    for i, h in enumerate(header, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        c.alignment = center
    r = 2
    for row in grid["rows"]:
        ws.cell(row=r, column=1, value=row["staff_num"])
        ws.cell(row=r, column=2, value=row["name"])
        # has_work=False は集計を出さず空欄のまま（legacy ゲートを踏襲）。
        if row["has_work"] and row["stats"]:
            for i, label in enumerate(stats_cols):
                ws.cell(row=r, column=3 + i, value=row["stats"][label]).alignment = center
        r += 1
    ws.column_dimensions["B"].width = 14


def _build_validation_sheet(wb, year, month, warnings):
    """検証レポートシート（live recompute_stats 由来の警告文字列を描画）。
    §6.5: 警告色は本体グリッドには出さず、この専用シートだけで着色する。"""
    warnings = warnings or []
    ws = wb.create_sheet("検証レポート(自動診断)")
    ws.cell(row=1, column=1,
            value=f"{year}年{month}月 勤務表 検証レポート").font = Font(size=14, bold=True)
    ws.merge_cells("A1:E1")

    if not warnings:
        ws.cell(row=3, column=1, value="状態: ✅ 正常").font = Font(color="0070C0", bold=True)
        ws.cell(row=5, column=1,
                value="すべての配置・スキル要件が正常に満たされています（問題なし）。")
    else:
        ws.cell(row=3, column=1,
                value=f"状態: ⚠️ {len(warnings)}件の警告あり").font = Font(color="FF0000", bold=True)
        ws.cell(row=5, column=1, value="【レポート詳細】").font = Font(bold=True)
        r = 6
        for w in warnings:
            cell = ws.cell(row=r, column=1, value=f"- {w}")
            if "不足" in w:
                cell.font = Font(color="C00000")       # 濃い赤
            elif "代替処理" in w:
                cell.font = Font(color="E36C0A")       # オレンジ
            r += 1
    ws.column_dimensions["A"].width = 100


def _apply_print_setup(ws, last_col):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.print_title_rows = "1:3"
    ws.print_title_cols = "A:B"
    ws.print_area = f"A1:{get_column_letter(last_col)}{ws.max_row}"
