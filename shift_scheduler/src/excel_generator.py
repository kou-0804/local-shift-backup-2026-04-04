"""
Excel出力生成器
"""
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from typing import Dict, List
from datetime import date
import calendar
from .models.staff import Staff

class ExcelGenerator:
    """勤務表Excel生成"""
    
    def __init__(
        self,
        year: int,
        month: int,
        technicians: List[Staff],
        night_assignments: Dict[int, List[str]],
        day_assignments: Dict[int, Dict[str, List[str]]],
        requests: Dict[int, Dict[str, str]], # {day: {staff_id: symbol}}
        on_call_assignments: Dict[int, Dict[str, str]] = None,
        name_mapper=None,
        daikyu_counts: Dict[str, int] = None,
        off_counts: Dict[str, int] = None,
        validation_errors: List[str] = None
    ):
        self.year = year
        self.month = month
        self.technicians = technicians
        self.night_assignments = night_assignments
        self.day_assignments = day_assignments
        self.requests = requests
        self.on_call_assignments = on_call_assignments or {}
        self.name_mapper = name_mapper
        self.daikyu_counts = daikyu_counts or {}
        self.off_counts = off_counts or {}
        self.validation_errors = validation_errors or []
        
        # 核医学・放射線治療のスタッフIDセット（表示制御用）
        self.nuc_tx_ids = {t.id for t in technicians if t.note and ('核医学' in t.note or '治療' in t.note)}
        
        self.days_in_month = calendar.monthrange(year, month)[1]
        
        # ワークブック作成
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = f"{month}月勤務表"
    
    def generate(self, output_path: str):
        """Excel生成"""
        self._create_header()
        self._create_day_header()
        self._create_weekday_row()
        self._fill_assignments()
        self._apply_formatting()
        
        if self.validation_errors is not None:
            self._create_validation_report_sheet()
        
        self.wb.save(output_path)
        print(f"✓ 勤務表を保存: {output_path}")
    
    def _create_header(self):
        """タイトル行"""
        self.ws['A1'] = f"画 像 診 断 室  {self.year}年 {self.month}月 勤 務 分 担 表"
        self.ws.merge_cells('A1:AH1')
        self.ws['A1'].alignment = Alignment(horizontal='center')
        self.ws['A1'].font = Font(size=14, bold=True)
    
    def _day_to_column(self, day: int) -> str:
        """日付を列名に変換（1→C, 2→D, ...）"""
        col_num = day + 2  # A=勤務表番号, B=技師名, C=1日
        return openpyxl.utils.get_column_letter(col_num)

    def _create_day_header(self):
        """日付ヘッダー"""
        self.ws['A2'] = '勤務表番号'
        self.ws['B2'] = '技師名'
        self.ws['B2'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.ws['B2'].font = Font(bold=True, color='FFFFFF')
        
        for day in range(1, self.days_in_month + 1):
            col = self._day_to_column(day)
            self.ws[f'{col}2'] = f'{day:02d}'
            self.ws[f'{col}2'].alignment = Alignment(horizontal='center')
            self.ws[f'{col}2'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            self.ws[f'{col}2'].font = Font(bold=True, color='FFFFFF')

        # Stats Header
        stats_start_col = self.days_in_month + 3
        self.stats_columns = ['夜勤', '病院MR', 'CLMR', '病CT', 'CT', 'ア', '心', 'ク', 'ポ', '精', 'MG', 'DR', 'HB', 'OP', '入', '病L', '超遅', 'ク遅', 'M遅', '公休', '代休']
        
        for i, label in enumerate(self.stats_columns):
            col_idx = stats_start_col + i
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            self.ws[f'{col_letter}2'] = label
            self.ws[f'{col_letter}2'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            self.ws[f'{col_letter}2'].font = Font(bold=True, color='FFFFFF')
            self.ws[f'{col_letter}2'].alignment = Alignment(horizontal='center')

    def _create_weekday_row(self):
        """曜日行"""
        self.ws['B3'] = '曜日'
        
        weekday_names = ['月', '火', '水', '木', '金', '土', '日']
        
        for day in range(1, self.days_in_month + 1):
            d = date(self.year, self.month, day)
            weekday = weekday_names[d.weekday()]
            
            col = self._day_to_column(day)
            self.ws[f'{col}3'] = weekday
            self.ws[f'{col}3'].alignment = Alignment(horizontal='center')


    def _fill_assignments(self):
        """配置を記入"""
        row = 4
        
        for tech in self.technicians:
            if tech.status != '在籍':
                continue
            
            # 技師番号・氏名
            try:
                tech_num = int(tech.id.replace('T', ''))
            except:
                tech_num = tech.id
                
            self.ws[f'A{row}'] = tech_num
            self.ws[f'B{row}'] = tech.name
            
            # カウンター初期化
            if not hasattr(self, 'stats_columns'):
                self.stats_columns = ['夜勤', '病院MR', 'CLMR', '病CT', 'CT', 'ア', '心', 'ク', 'ポ', '精', 'MG', 'DR', 'HB', 'OP', '入', '病L', '超遅', 'ク遅', 'M遅', '公休', '代休']
            counts = {label: 0 for label in self.stats_columns}
            # 実勤務（勤務地アサイン）が1件も無い行＝当月スケジュール対象外・全休・育休等は
            # 集計を出さない（空白セルが全日「休」と数えられ公休=月日数になる無意味な値を避ける）。
            # 勤務地コード = 撮影部の業務 + 外部部門(PET/RI/放治/TV/館山)。
            # 休・☆・育休・出・講・全会・研 等は勤務地アサインに含めない。
            WORK_LOCATION_CODES = {
                '病院MR', 'CLMR', 'CT', '病CT', 'ア', '心', 'ク', 'クL', 'ポ', '精',
                'MG', 'DR', 'HB', 'OP', 'PICC', '入', '病L', '超遅', 'ク遅', 'M遅',
                '館山', 'TV', 'PET', 'RI', '放治', 'DX',
            }
            row_has_work = False

            # 各日の配置
            for day in range(1, self.days_in_month + 1):
                col = self._day_to_column(day)
                cell_value = self._get_assignment_text(tech.id, day)
                self.ws[f'{col}{row}'] = cell_value
                self.ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
                
                # 色分け
                fill = self._get_cell_fill(tech.id, day, cell_value)
                if fill:
                    self.ws[f'{col}{row}'].fill = fill
                
                # 統計カウント
                if '夜' in cell_value:
                    counts['夜勤'] += 1

                # 日勤カウント (cell_value might be "CT", "CT/夜", "入")
                # Split by '/' if composite
                # さらに「日勤＋夜勤」が "病CT夜"・"CLMR夜(希)" のようにスラッシュ無しで
                # 連結されるケースは、末尾の "(希)"・"夜" を剥がして業務名へ正規化してから
                # 一致させ、日勤業務としても計上する（夜勤列は上の '夜' 判定で別途 +1 済み）。
                parts = cell_value.split('/')
                for p in parts:
                    p = p.strip()
                    if p.endswith('(希)'):
                        p = p[:-3]
                    elif p.endswith('（希）'):
                        p = p[:-3]
                    if p.endswith('夜'):
                        p = p[:-1]
                    if p in WORK_LOCATION_CODES:
                        row_has_work = True
                    if p == 'クL':   # クL(クリニックリーダー)はクの内数として集計
                        p = 'ク'
                    if p in counts:
                        counts[p] += 1
            
            # 公休・代休カウント（assign_monthly_off_days の計算結果を使用）
            counts['公休'] = self.off_counts.get(tech.id, 0)
            counts['代休'] = self.daikyu_counts.get(tech.id, 0)
            
            # 統計出力（実勤務が無い行は、空白セルが全日「休」と数えられ公休=月日数 等の
            # 無意味な値になるため、集計列を出さず空欄のままにする）
            if row_has_work:
                stats_start_col = self.days_in_month + 3
                for i, label in enumerate(self.stats_columns):
                    col_idx = stats_start_col + i
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    self.ws[f'{col_letter}{row}'] = counts[label]
                    self.ws[f'{col_letter}{row}'].alignment = Alignment(horizontal='center')

            row += 1

        # Add On-Call rows
        if self.on_call_assignments:
            oncall_labels = ['第1拘束', '第2拘束']
            staff_dict = {t.id: t.name for t in self.technicians}
            
            for label in oncall_labels:
                self.ws[f'A{row}'] = ''
                self.ws[f'B{row}'] = label
                
                for day in range(1, self.days_in_month + 1):
                    col = self._day_to_column(day)
                    
                    assigned_staff_id = self.on_call_assignments.get(day, {}).get(label)
                    if assigned_staff_id:
                        # Extract lastname for shorter display
                        name = staff_dict.get(assigned_staff_id, assigned_staff_id)
                        # Name usually looks like '佐藤(海)'. User's template shows '佐藤海'.
                        # I'll strip parentheses for oncall display
                        name = name.replace('(', '').replace(')', '').replace(' ', '').replace('　', '')
                        self.ws[f'{col}{row}'] = name
                    self.ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
                    
                row += 1

    def _get_assignment_text(self, tech_id: str, day: int) -> str:
        """配置テキストを取得"""
        parts = []
        
        # 日勤配置
        if day in self.day_assignments:
            for loc_code, tech_ids in self.day_assignments[day].items():
                if tech_id in tech_ids:
                    parts.append(loc_code)
        
        # 夜勤
        if day in self.night_assignments:
            if tech_id in self.night_assignments[day]:
                if parts:
                    parts[0] += '夜'
                else:
                    parts.append('夜')
        
        # 明け
        if day > 1 and (day - 1) in self.night_assignments:
            if tech_id in self.night_assignments[day - 1]:
                return '○'
        
        # Visualizing Requests (User Requirement)
        req_symbol = ""
        if day in self.requests and tech_id in self.requests[day]:
             req_symbol = self.requests[day][tech_id]
             
        # If Night Request matches Assignment, append (希)
        if req_symbol == '夜希':
             # Find the part with '夜' and mark it
             for i, p in enumerate(parts):
                  if '夜' in p:
                      parts[i] = p + '(希)'

        # Fix: Ensure 17業/17休 is reflected
        if req_symbol in ['17業', '17休']:
            # append to list effectively
            parts.append(req_symbol)
        
        # 割り当てがない場合、申請を表示
        if not parts:
            if req_symbol and req_symbol != '休(仮)':
                return req_symbol
        
        # 核医学・治療のスタッフで、システムが付けた「休」を非表示にする
        # （夜勤・明け・申請休みがある場合はそちらが優先されるため、ここでは純粋な「休」を空欄化する）
        if tech_id in self.nuc_tx_ids and parts == ['休']:
            # 申請が休み系（◆, ☆, 17休など）でない場合は、空欄にする
            if not req_symbol or req_symbol in ['', '夜希', '休(仮)']:
                return ''

        # 「休」が割り当てられていても、予定申請があればそちらを優先表示
        if parts == ['休'] and req_symbol and req_symbol not in ['', '夜希', '休(仮)']:
            return req_symbol
        
        return '/'.join(parts) if parts else ''
    
    def _get_cell_fill(self, tech_id: str, day: int, cell_value: str) -> PatternFill:
        """セルの背景色を取得"""
        if '夜' in cell_value:
            return PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        if cell_value == '○':
            return PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')
        if cell_value in ['★', '☆', '◆']:
            return PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
        if cell_value == '休':
            # Enforced Holiday (Grey)
            return PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        return None

    def _apply_formatting(self):
        """書式設定"""
        # 罫線
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        if not hasattr(self, 'stats_columns'):
             self.stats_columns = ['夜勤', '病院MR', 'CLMR', '病CT', 'CT', 'ア', '心', 'ク', 'ポ', '精', 'MG', 'DR', 'HB', 'OP', '入', '病L', '超遅', 'ク遅', 'M遅', '公休', '代休']
             
        max_col = self.days_in_month + 2 + len(self.stats_columns) # +2 for ID, Name
        
        for row in self.ws.iter_rows(min_row=2, max_row=self.ws.max_row,
                                       min_col=1, max_col=max_col):
            for cell in row:
                cell.border = thin_border
        
        # 列幅調整
        self.ws.column_dimensions['A'].width = 8
        self.ws.column_dimensions['B'].width = 15
        for day in range(1, self.days_in_month + 1):
            col = self._day_to_column(day)
            self.ws.column_dimensions[col].width = 4 # Narrow for days
            
        # Stats width
        stats_start_col = self.days_in_month + 3
        for i in range(len(self.stats_columns)):
             col_letter = openpyxl.utils.get_column_letter(stats_start_col + i)
             self.ws.column_dimensions[col_letter].width = 4

    def _create_validation_report_sheet(self):
        """検証レポートシートを作成"""
        ws_val = self.wb.create_sheet('検証レポート(自動診断)')
        
        ws_val['A1'] = f"{self.year}年{self.month}月 勤務表 検証レポート"
        ws_val.merge_cells('A1:E1')
        ws_val['A1'].font = Font(size=14, bold=True)
        
        if not self.validation_errors:
            ws_val['A3'] = "状態: ✅ 正常"
            ws_val['A3'].font = Font(color='0070C0', bold=True) # 濃い青/緑系の代替
            ws_val['A5'] = "すべての配置・スキル要件が正常に満たされています。"
        else:
            ws_val['A3'] = f"状態: ⚠️ {len(self.validation_errors)}件の警告あり"
            ws_val['A3'].font = Font(color='FF0000', bold=True)
            
            ws_val['A5'] = "【レポート詳細】"
            ws_val['A5'].font = Font(bold=True)
            
            row = 6
            for error in self.validation_errors:
                ws_val[f'A{row}'] = f"- {error}"
                
                # エラーの種類によって文字色を分ける
                if "不足" in error:
                    ws_val[f'A{row}'].font = Font(color='C00000') # 濃い赤
                elif "代替処理" in error:
                    ws_val[f'A{row}'].font = Font(color='E36C0A') # オレンジ
                    
                row += 1
                
        ws_val.column_dimensions['A'].width = 100
