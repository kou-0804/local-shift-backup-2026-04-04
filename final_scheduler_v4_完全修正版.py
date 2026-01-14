"""
勤務表作成システム 最終版 v4.0
全要件完全対応・汎用化対応済み
"""
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
import pandas as pd
import json
from datetime import datetime, date
from collections import defaultdict
import jpholiday
import calendar

print("=" * 70)
print("勤務表作成システム 最終版 v4.0")
print("=" * 70)
print()

# 年月の入力
TARGET_YEAR = int(input("年を入力 (例: 2025): "))
TARGET_MONTH = int(input("月を入力 (例: 12): "))
print()

# 日数と祝日の自動計算
last_day = calendar.monthrange(TARGET_YEAR, TARGET_MONTH)[1]
holidays_list = jpholiday.between(
    date(TARGET_YEAR, TARGET_MONTH, 1),
    date(TARGET_YEAR, TARGET_MONTH, last_day)
)
HOLIDAYS_TARGET = [d[0].day for d in holidays_list]

print(f"📅 {TARGET_YEAR}年{TARGET_MONTH}月の勤務表を作成します")
print(f"   日数: {last_day}日")
print(f"   祝日: {HOLIDAYS_TARGET}")
print()

# ============================================
# 設定
# ============================================

EXCLUDED_STAFF = ['友邉　和哉', '松本　梓']
AG_TUESDAY = ['清水　万慈', '児玉　勇輝', '松本　梓', '佐藤　和彦', '池谷　尚人']
SEIMITSU_WED_FRI = ['須田　章則', '永井　基博', '石川　和弥']
EXPERIENCE_THRESHOLD = 3
MAX_CONSECUTIVE_DAYS = 6

# ============================================
# データ読み込み
# ============================================

print("📂 データ読み込み...")

df_skills = pd.read_csv('技師スキルマスタ_完全版.csv')
staff_order = df_skills['技師名'].tolist()
skill_matrix = {}
gender_map = {}
experience_map = {}

for _, row in df_skills.iterrows():
    name = row['技師名']
    gender_map[name] = row.get('性別', '')
    experience_map[name] = int(row.get('経験年数', 0))
    skills = {}
    for col in df_skills.columns:
        if col not in ['技師名', '性別', '経験年数']:
            skills[col] = (row[col] == '○')
    skill_matrix[name] = skills

with open(f'予定申請_分類済み_{TARGET_MONTH}月.json', 'r', encoding='utf-8') as f:
    vacation_data = json.load(f)

gyohai_days = []
for name, data in vacation_data.items():
    if 'gyohai' in data:
        gyohai_days.extend(data['gyohai'])
gyohai_days = list(set(gyohai_days))

df_night = pd.read_csv(f'夜勤割り当て_{TARGET_MONTH}月.csv')

df_locations = pd.read_csv('勤務場所_設定.csv')
location_needs = {}
for _, row in df_locations.iterrows():
    loc = row['勤務場所']
    location_needs[loc] = {
        '月': int(row.get('月', 1)),
        '火': int(row.get('火', 1)),
        '水': int(row.get('水', 1)),
        '木': int(row.get('木', 1)),
        '金': int(row.get('金', 1)),
        '土': int(row.get('土', 0)),
        '日': int(row.get('日', 0))
    }

print(f"  技師: {len(staff_order)}名")
print(f"  業配: {len(gyohai_days)}日")
print()

# ============================================
# Excel作成
# ============================================

print("📊 Excel作成...")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = f"{TARGET_MONTH}月"

yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
green_night = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
orange = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
pink = PatternFill(start_color='FF6B9D', end_color='FF6B9D', fill_type='solid')
light_yellow = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
light_purple = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
green = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
dark_gray = PatternFill(start_color='404040', end_color='404040', fill_type='solid')  # 追加
light_gray = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')  # 追加
white_font = Font(color='FFFFFF')  # 追加
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')

ws['A1'] = f'{TARGET_YEAR}年{TARGET_MONTH}月 勤務表'
ws['A1'].font = Font(size=14, bold=True)

ws['B2'] = '技師名'
ws['B2'].fill = header_fill
ws['B2'].font = header_font

WEEKDAYS = ['月', '火', '水', '木', '金', '土', '日']
for day in range(1, last_day + 1):
    col = day + 2
    dt = datetime(TARGET_YEAR, TARGET_MONTH, day)
    weekday = WEEKDAYS[dt.weekday()]
    ws.cell(2, col, f"{day}\n{weekday}")
    ws.cell(2, col).fill = header_fill
    ws.cell(2, col).font = header_font
    ws.cell(2, col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

stats_col = last_day + 3
for i, label in enumerate(['夜勤', '超遅', 'MG', 'ク', '遅番', '病CT', 'CT', '入', 'ポ', '精', 'HB', 'OP', '心', 'ア', 'DR']):
    ws.cell(2, stats_col + i, label).fill = header_fill
    ws.cell(2, stats_col + i).font = header_font

for idx, name in enumerate(staff_order, start=3):
    ws.cell(idx, 2, name)

# ============================================
# 夜勤配置
# ============================================

print("🌙 夜勤配置...")

night_count = defaultdict(int)
holiday_work_count = {name: 0 for name in staff_order}  # 日祝日勤務カウンター

for _, row in df_night.iterrows():
    date_str = row['日付']
    day = int(date_str.split('/')[2])
    
    for i in range(1, 4):
        staff_name = row[f'夜勤{i}']
        if pd.notna(staff_name) and staff_name in staff_order:
            row_idx = staff_order.index(staff_name) + 3
            
            is_night_request = False
            if staff_name in vacation_data:
                if day in vacation_data[staff_name].get('night_request', []):
                    is_night_request = True
            
            cell = ws.cell(row_idx, day + 2)
            cell.value = '夜'
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = green_night if is_night_request else yellow
            
            night_count[staff_name] += 1
            
            # 翌日が日祝日なら夜勤者にカウント
            if day < last_day:
                next_dt = datetime(TARGET_YEAR, TARGET_MONTH, day + 1)
                next_weekday = next_dt.weekday()
                if next_weekday == 6 or (day + 1) in HOLIDAYS_TARGET:
                    holiday_work_count[staff_name] += 1
            
            if day < last_day:
                ws.cell(row_idx, day + 3, '○').fill = orange
                ws.cell(row_idx, day + 3).alignment = Alignment(horizontal='center', vertical='center')

print("  夜勤配置完了")

# ============================================
# 休み希望配置
# ============================================

print("🌟 休み希望配置...")

for name, data in vacation_data.items():
    if name not in staff_order:
        continue
    
    row_idx = staff_order.index(name) + 3
    
    # 休み希望（★）
    for day in data.get('star_yasumi', []):
        if 1 <= day <= last_day:
            cell = ws.cell(row_idx, day + 2)
            if not cell.value:
                cell.value = '★'
                cell.fill = pink
                cell.alignment = Alignment(horizontal='center', vertical='center')

print(f"  休み希望配置完了")

# ============================================
# その他の予定記号配置
# ============================================

print("📝 予定記号配置...")
other_mark_count = 0

for name, data in vacation_data.items():
    if name not in staff_order:
        continue
    
    row_idx = staff_order.index(name) + 3
    
    # その他の記号を配置
    for day_str, symbol in data.get('other_marks', {}).items():
        day = int(day_str)
        if 1 <= day <= last_day:
            cell = ws.cell(row_idx, day + 2)
            if not cell.value:
                cell.value = symbol
                cell.alignment = Alignment(horizontal='center', vertical='center')
                other_mark_count += 1

print(f"  その他記号配置完了: {other_mark_count}件")

# ============================================
# 夜勤明け後の休み
# ============================================

print("💫 夜勤明け後の休み調整...")

for day in range(1, last_day):
    for name in staff_order:
        row_idx = staff_order.index(name) + 3
        today_cell = ws.cell(row_idx, day + 2).value
        next_cell = ws.cell(row_idx, day + 3)
        
        if today_cell == '○' and not next_cell.value:
            is_day_off_request = False
            if name in vacation_data:
                if (day + 1) in vacation_data[name].get('star_yasumi', []):
                    is_day_off_request = True
            
            next_cell.value = '★' if is_day_off_request else '☆'
            next_cell.fill = pink if is_day_off_request else light_yellow
            next_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 明けの日が日祝日ならカウント
            next_dt = datetime(TARGET_YEAR, TARGET_MONTH, day + 1)
            next_weekday = next_dt.weekday()
            if next_weekday == 6 or (day + 1) in HOLIDAYS_TARGET:
                holiday_work_count[name] += 1

print("  完了")

# ============================================
# 連勤カウント
# ============================================

def count_consecutive_days(staff_name, day):
    if day == 1:
        return 0
    
    row_idx = staff_order.index(staff_name) + 3
    count = 0
    
    for d in range(day - 1, 0, -1):
        cell_value = ws.cell(row_idx, d + 2).value
        if cell_value and cell_value not in ['★', '☆']:
            count += 1
        else:
            break
    
    return count

# ============================================
# 日勤配置
# ============================================

print("☀️ 日勤配置...")

load_score = {name: 0 for name in staff_order}
location_count = {name: defaultdict(int) for name in staff_order}
choren_count = {name: 0 for name in staff_order}
mg_count = {name: 0 for name in staff_order}
soban_count = {name: 0 for name in staff_order}
ku_count = {name: 0 for name in staff_order}
last_location = {name: None for name in staff_order}

for day in range(1, last_day + 1):
    dt = datetime(TARGET_YEAR, TARGET_MONTH, day)
    weekday_num = dt.weekday()
    weekday = WEEKDAYS[weekday_num]
    
    is_third_saturday = (weekday_num == 5 and 15 <= day <= 21)
    is_sunday_holiday = (weekday_num == 6 or day in HOLIDAYS_TARGET)
    is_gyohai = (day in gyohai_days)
    is_first_friday = (weekday_num == 4 and 1 <= day <= 7)
    is_fourth_thursday = (weekday_num == 3 and 22 <= day <= 28)
    
    print(f"  {day}日({weekday})...", end=" ")
    
    if is_third_saturday:
        print("休診")
        continue
    
    assigned = []
    
    available = []
    has_night_shift = {}  # 夜勤がある人を記録
    has_17_mark = {}  # 17業・17休がある人を記録
    original_marks = {}  # 元の記号を記録
    
    for name in staff_order:
        if name in EXCLUDED_STAFF:
            continue
        
        row_idx = staff_order.index(name) + 3
        cell_value = ws.cell(row_idx, day + 2).value
        
        # 配置不可の記号
        if cell_value in ['○', '★', '☆', '出', '講', '会議', '全会']:
            continue
        
        # 夜勤がある場合は記録して配置可能
        if cell_value == '夜':
            has_night_shift[name] = True
            original_marks[name] = '夜'
        
        # 17業・17休がある場合は記録して配置可能
        if cell_value in ['17業', '17休']:
            has_17_mark[name] = True
            original_marks[name] = cell_value
        
        # 業配・業出がある場合
        if cell_value in ['業配', '業出']:
            original_marks[name] = cell_value
            if gender_map.get(name) == '女':
                has_17_mark[name] = True  # クに配置するため記録
            elif name == '川名　佑樹':
                has_17_mark[name] = True  # MRに配置するため記録
            else:
                continue  # その他男性は配置しない
        
        consecutive = count_consecutive_days(name, day)
        if consecutive >= MAX_CONSECUTIVE_DAYS:
            continue
        
        available.append(name)
    
    if is_sunday_holiday:
        candidates = [n for n in available if skill_matrix.get(n, {}).get('出', False)]
        # 日祝日勤務回数が少ない人を優先
        candidates.sort(key=lambda x: (holiday_work_count[x], load_score[x]))
        
        for name in candidates[:2]:
            row_idx = staff_order.index(name) + 3
            ws.cell(row_idx, day + 2, '出').fill = green
            ws.cell(row_idx, day + 2).alignment = Alignment(horizontal='center', vertical='center')
            load_score[name] += 1
            holiday_work_count[name] += 1
        print("出勤2名")
        continue
    
    # 業配・業出の特別配置
    for name in staff_order:
        row_idx = staff_order.index(name) + 3
        cell_value = ws.cell(row_idx, day + 2).value
        
        if cell_value in ['業配', '業出']:
            cell = ws.cell(row_idx, day + 2)
            if gender_map.get(name) == '女':
                # 女性 → ク/業配
                cell.value = f'ク/{cell_value}'
                cell.fill = light_gray
                cell.alignment = Alignment(horizontal='center', vertical='center')
                assigned.append(name)
                load_score[name] += 1
                location_count[name]['ク'] += 1
                ku_count[name] += 1
            elif name == '川名　佑樹':
                # 川名 → MR/業配
                cell.value = f'MR/{cell_value}'
                cell.fill = light_gray
                cell.alignment = Alignment(horizontal='center', vertical='center')
                assigned.append(name)
                load_score[name] += 1
    
    # HB TAVI
    if is_first_friday or is_fourth_thursday:
        tavi_candidates = [n for n in available if skill_matrix.get(n, {}).get('TAVI', False)]
        # HB回数が少ない人を優先
        tavi_candidates.sort(key=lambda x: (location_count[x]['HB'], load_score[x]))
        
        for name in tavi_candidates[:3]:
            row_idx = staff_order.index(name) + 3
            ws.cell(row_idx, day + 2, 'HB')
            ws.cell(row_idx, day + 2).alignment = Alignment(horizontal='center', vertical='center')
            assigned.append(name)
            load_score[name] += 1
            location_count[name]['HB'] += 1
        
        available = [n for n in available if n not in assigned]
    
    # 病CT優先配置
    byoct_needed = location_needs.get('病CT', {}).get(weekday, 0)
    if byoct_needed > 0:
        byoct_candidates = [n for n in available 
                           if n not in assigned 
                           and skill_matrix.get(n, {}).get('病CT', False)]
        
        shinzo_ct = [n for n in byoct_candidates if skill_matrix.get(n, {}).get('心CT', False)]
        other_byoct = [n for n in byoct_candidates if n not in shinzo_ct]
        
        # 病CT回数が少ない人を優先、その中で全体負荷が低い人を選ぶ
        shinzo_ct.sort(key=lambda x: (location_count[x]['病CT'], load_score[x]))
        selected_shinzo = shinzo_ct[:min(3, len(shinzo_ct))]
        
        other_byoct.sort(key=lambda x: (location_count[x]['病CT'], load_score[x]))
        remaining_needed = byoct_needed - len(selected_shinzo)
        selected_other = other_byoct[:remaining_needed]
        
        for name in selected_shinzo + selected_other:
            row_idx = staff_order.index(name) + 3
            ws.cell(row_idx, day + 2, '病CT')
            ws.cell(row_idx, day + 2).alignment = Alignment(horizontal='center', vertical='center')
            assigned.append(name)
            load_score[name] += 1
            location_count[name]['病CT'] += 1
    
    # 通常日勤配置
    for location in location_needs.keys():
        if location == '病CT':
            continue
        
        needed = location_needs[location].get(weekday, 0)
        if needed == 0:
            continue
        
        if location == 'ア' and weekday == '火':
            candidates = [n for n in available if n in AG_TUESDAY and n not in assigned]
            needed = 2
        elif location == '精':
            # 精密検査の配置
            if weekday in ['水', '金']:
                # 水曜・金曜は須田、永井、石川のいずれか（回数が少ない順）
                candidates = [n for n in available if n in SEIMITSU_WED_FRI and n not in assigned]
                candidates.sort(key=lambda x: (location_count[x]['精'], load_score[x]))
            else:
                # 月・火・木・土は精スキルがある人全員から均等に（回数が少ない順）
                candidates = [n for n in available 
                             if n not in assigned 
                             and skill_matrix.get(n, {}).get('精', False)]
                candidates.sort(key=lambda x: (location_count[x]['精'], load_score[x]))
        elif location.startswith('M') and is_gyohai:
            candidates = [n for n in available 
                         if n not in assigned 
                         and skill_matrix.get(n, {}).get(location, False)]
            if '川名　佑樹' in candidates:
                candidates = ['川名　佑樹'] + [c for c in candidates if c != '川名　佑樹']
        elif location == 'ク' and is_gyohai:
            # 女性を最低2名確保
            female_candidates = [n for n in available 
                               if n not in assigned 
                               and skill_matrix.get(n, {}).get(location, False)
                               and gender_map.get(n, '') == '女']
            male_candidates = [n for n in available 
                             if n not in assigned 
                             and skill_matrix.get(n, {}).get(location, False)
                             and gender_map.get(n, '') == '男']
            
            # クの回数が少ない人を優先
            female_candidates.sort(key=lambda x: (location_count[x][location], load_score[x]))
            selected = female_candidates[:min(2, len(female_candidates))]
            
            remaining = female_candidates[len(selected):] + male_candidates
            remaining.sort(key=lambda x: (location_count[x][location], load_score[x]))
            candidates = selected + remaining
        elif location == '超遅':
            # 超遅の制約
            yesterday_choren = []
            if day > 1:
                for name in staff_order:
                    row_idx = staff_order.index(name) + 3
                    if ws.cell(row_idx, day + 1).value == '超遅':
                        yesterday_choren.append(name)
            
            tomorrow_night = []
            if day < last_day:
                tomorrow_row = df_night[df_night['日付'] == f'{TARGET_YEAR}/{TARGET_MONTH:02d}/{day+1:02d}']
                if not tomorrow_row.empty:
                    tomorrow_night = [
                        tomorrow_row.iloc[0].get('夜勤1'),
                        tomorrow_row.iloc[0].get('夜勤2'),
                        tomorrow_row.iloc[0].get('夜勤3')
                    ]
                    tomorrow_night = [n for n in tomorrow_night if pd.notna(n)]
            
            # 候補者を選出
            male_candidates = []
            female_with_mg = []
            female_without_mg = []
            
            for n in available:
                if n in assigned or n in yesterday_choren or n in tomorrow_night:
                    continue
                if n in has_night_shift:  # 夜勤がある人を除外
                    continue
                if not skill_matrix.get(n, {}).get('超遅', False):
                    continue
                
                gender = gender_map.get(n, '')
                if gender == '男':
                    male_candidates.append(n)
                elif gender == '女':
                    if skill_matrix.get(n, {}).get('MG', False):
                        female_with_mg.append(n)
                    else:
                        female_without_mg.append(n)
            
            # 超遅回数で均等化
            prioritized = male_candidates + female_with_mg
            prioritized.sort(key=lambda x: (choren_count.get(x, 0), load_score[x]))
            
            female_without_mg.sort(key=lambda x: (choren_count.get(x, 0), load_score[x]))
            
            candidates = prioritized + female_without_mg
        elif location == 'MG':
            # MGの制約と優先順位
            yesterday_choren_female = []
            if day > 1:
                for name in staff_order:
                    row_idx = staff_order.index(name) + 3
                    if ws.cell(row_idx, day + 1).value == '超遅' and gender_map.get(name) == '女':
                        yesterday_choren_female.append(name)
            
            # 前日超遅の女性（MGスキルある人のみ）
            priority_candidates = [n for n in yesterday_choren_female
                                  if n in available 
                                  and n not in assigned
                                  and n not in has_night_shift  # 夜勤がある人を除外
                                  and skill_matrix.get(n, {}).get('MG', False)]
            
            # その他のMGスキル保有女性
            other_candidates = [n for n in available
                              if n not in assigned
                              and n not in priority_candidates
                              and n not in has_night_shift  # 夜勤がある人を除外
                              and skill_matrix.get(n, {}).get('MG', False)
                              and gender_map.get(n) == '女']
            
            # 均等化のため、mg_count少ない順にソート
            priority_candidates.sort(key=lambda x: (mg_count.get(x, 0), load_score[x]))
            other_candidates.sort(key=lambda x: (mg_count.get(x, 0), load_score[x]))
            
            candidates = priority_candidates + other_candidates
        elif location == 'ポ':
            # 経験年数による上限と均等化
            candidates_junior = []
            candidates_senior = []
            
            for n in available:
                if n in assigned:
                    continue
                if n in has_night_shift:  # 夜勤がある人を除外
                    continue
                if not skill_matrix.get(n, {}).get(location, False):
                    continue
                
                exp = experience_map.get(n, 0)
                po_count = location_count[n]['ポ']
                
                if exp <= 6:
                    if po_count < 3:
                        candidates_junior.append(n)
                else:
                    if po_count < 1:
                        candidates_senior.append(n)
            
            candidates_junior.sort(key=lambda x: (location_count[x]['ポ'], load_score[x]))
            candidates_senior.sort(key=lambda x: (location_count[x]['ポ'], load_score[x]))
            
            candidates = candidates_senior + candidates_junior
        elif location == 'ク':
            # 通常日も女性最低2名
            female_candidates = [n for n in available 
                               if n not in assigned 
                               and skill_matrix.get(n, {}).get(location, False)
                               and gender_map.get(n, '') == '女']
            male_candidates = [n for n in available 
                             if n not in assigned 
                             and skill_matrix.get(n, {}).get(location, False)
                             and gender_map.get(n, '') == '男']
            
            # クの回数が少ない人を優先
            female_candidates.sort(key=lambda x: (location_count[x][location], load_score[x]))
            selected = female_candidates[:min(2, len(female_candidates))]
            
            remaining = female_candidates[len(selected):] + male_candidates
            remaining.sort(key=lambda x: (location_count[x][location], load_score[x]))
            candidates = selected + remaining
        else:
            candidates = [n for n in available 
                         if n not in assigned 
                         and skill_matrix.get(n, {}).get(location, False)]
            
            # 優先順位
            candidates.sort(key=lambda x: (
                location_count[x][location],
                10 if last_location.get(x) == location else 0,
                load_score[x]
            ))            
        
        for name in candidates[:needed]:
            row_idx = staff_order.index(name) + 3
            cell = ws.cell(row_idx, day + 2)
            
            # 元の記号がある場合は併記
            if name in original_marks:
                cell.value = f"{location}/{original_marks[name]}"
            else:
                cell.value = location
            
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 色設定
            if name in has_night_shift:
                cell.fill = dark_gray
                cell.font = white_font
            elif name in has_17_mark:
                cell.fill = light_gray
            
            assigned.append(name)
            load_score[name] += 1
            location_count[name][location] += 1
            last_location[name] = location
            
            if location == '超遅':
                choren_count[name] += 1
            elif location == 'MG':
                mg_count[name] += 1
            elif location == 'ク':
                ku_count[name] += 1
    
    print(f"{len(assigned)}人")

# ============================================
# 超遅の翌日
# ============================================

print("🔧 超遅の翌日調整...")
for day in range(1, last_day):
    for name in staff_order:
        row_idx = staff_order.index(name) + 3
        today = ws.cell(row_idx, day + 2).value
        
        if today == '超遅':
            next_cell = ws.cell(row_idx, day + 3)
            gender = gender_map.get(name, '')
            
            if next_cell.value:
                continue
            
            # 休み希望がある場合
            is_day_off_request = False
            if name in vacation_data:
                if (day + 1) in vacation_data[name].get('star_yasumi', []):
                    is_day_off_request = True
            
            if is_day_off_request:
                next_cell.value = '★'
                next_cell.fill = pink
            elif gender == '女':
                # 女性: MGに配置（MGスキルがある場合）
                if skill_matrix.get(name, {}).get('MG', False):
                    next_cell.value = 'MG'
                    next_cell.alignment = Alignment(horizontal='center', vertical='center')
                    mg_count[name] += 1
                else:
                    next_cell.value = '☆'
                    next_cell.fill = light_purple
            else:
                next_cell.value = '☆'
                next_cell.fill = light_purple

print("✅ 完了")

# ============================================
# 遅番の選出
# ============================================

print("🔧 遅番の選出...")

soban_row = len(staff_order) + 3 + 1
ws.cell(soban_row, 1, "遅番")
ws.cell(soban_row, 1).font = header_font
ws.cell(soban_row, 1).alignment = Alignment(horizontal='center', vertical='center')

soban_locations = ['入', '病CT', 'OP', '精', 'CT']

for day in range(1, last_day + 1):
    dt = datetime(TARGET_YEAR, TARGET_MONTH, day)
    weekday = dt.weekday()
    
    # 土日祝日はスキップ
    if weekday >= 5 or day in HOLIDAYS_TARGET:
        ws.cell(soban_row, day + 2, '-')
        ws.cell(soban_row, day + 2).alignment = Alignment(horizontal='center', vertical='center')
        continue
    
    # その日の入・病CT・OP・精・CTに配置された人を集める
    candidates = []
    for name in staff_order:
        row_idx = staff_order.index(name) + 3
        location = ws.cell(row_idx, day + 2).value
        
        # 夜勤チェック（「CT/夜」などを除外）
        if location and '夜' in str(location):
            continue
        
        if location in soban_locations:
            if skill_matrix.get(name, {}).get('出', False):
                candidates.append(name)
    
    candidates.sort(key=lambda x: soban_count.get(x, 0))
    
    if candidates:
        selected = candidates[0]
        ws.cell(soban_row, day + 2, selected)
        ws.cell(soban_row, day + 2).alignment = Alignment(horizontal='center', vertical='center')
        soban_count[selected] += 1

print("✅ 完了")

# ============================================
# 統計
# ============================================

for idx, name in enumerate(staff_order, start=3):
    ws.cell(idx, stats_col, night_count.get(name, 0))
    ws.cell(idx, stats_col + 1, choren_count.get(name, 0))
    ws.cell(idx, stats_col + 2, mg_count.get(name, 0))
    ws.cell(idx, stats_col + 3, ku_count.get(name, 0))
    ws.cell(idx, stats_col + 4, soban_count.get(name, 0))
    ws.cell(idx, stats_col + 5, location_count[name]['病CT'])
    ws.cell(idx, stats_col + 6, location_count[name]['CT'])
    ws.cell(idx, stats_col + 7, location_count[name]['入'])
    ws.cell(idx, stats_col + 8, location_count[name]['ポ'])
    ws.cell(idx, stats_col + 9, location_count[name]['精'])
    ws.cell(idx, stats_col + 10, location_count[name]['HB'])
    ws.cell(idx, stats_col + 11, location_count[name]['OP'])
    ws.cell(idx, stats_col + 12, location_count[name]['心'])
    ws.cell(idx, stats_col + 13, location_count[name]['ア'])
    ws.cell(idx, stats_col + 14, location_count[name]['DR'])

ws.column_dimensions['B'].width = 15
for col in range(3, last_day + 20):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 4

# ============================================
# 保存
# ============================================

output_file = f'勤務表_{TARGET_YEAR}年{TARGET_MONTH}月_最終版v4.xlsx'
wb.save(output_file)
print()
print("=" * 70)
print(f"✅ 完成: {output_file}")
print("=" * 70)

import subprocess
try:
    subprocess.run(['open', output_file])
except:
    pass
